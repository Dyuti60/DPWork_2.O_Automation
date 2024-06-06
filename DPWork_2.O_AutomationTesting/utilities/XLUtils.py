import openpyxl
import sys
from exception import CustomException
def createNewExcelFile(filepath):
    try:
        wb=openpyxl.Workbook()
        wb.save(filepath)
        wb.close()
    except Exception as e:
        raise CustomException(e,sys)

def renameSheet(oldSheetName, newSheetName, excelFilePath):
    try:
        sheet=openpyxl.load_workbook(excelFilePath)
        sheet_s=sheet[oldSheetName]
        sheet_s.title=newSheetName
        sheet.save(excelFilePath)
    except Exception as e:
        raise CustomException(e,sys)

def createSheet(file, sheet):
    try:
        workbook=openpyxl.load_workbook(file)
        workbook.create_sheet(sheet)
        workbook.save(file)
        workbook.close()
    except Exception as e:
        raise CustomException(e,sys)

def getRowCount(file, sheetName):
    try:
        workbook=openpyxl.load_workbook(file, sheetName)
        sheet=workbook[sheetName]
        count =(len([row for row in sheet if any(cell.value is not None for cell in row)]))
        return count
        #return sheet.max_row
    except Exception as e:
        raise CustomException(e,sys)

def getColCount(file, sheetName):
    try:
        workbook=openpyxl.load_workbook(file, sheetName)
        sheet=workbook[sheetName]
        return sheet.max_column
    except Exception as e:
        raise CustomException(e,sys)

def readData(file, sheetName, rownum, colnum):
    try:
        workbook=openpyxl.load_workbook(file, sheetName)
        sheet=workbook[sheetName]
        return sheet.cell(row=rownum, column=colnum).value
    except Exception as e:
        raise CustomException(e,sys)

def writeData(file, sheetName, rownum, colnum, data):
    try:
        workbook=openpyxl.load_workbook(file)
        sheet=workbook[sheetName]
        sheet.cell(rownum,colnum).value=data
        workbook.save(file)
        workbook.close()
    except Exception as e:
        raise CustomException(e,sys)

def getDataInList(file, sheetName, row, col):
    try:
        workbook=openpyxl.load_workbook(file)
        sheet=workbook[sheetName]

        data_list=[]

        for j in range(2, row+1):
            data=sheet.cell(j,col).value
            if data is not None:
                data_list.append(data)
            else:
                data_list.append('')
        return data_list
    except Exception as e:
        raise CustomException(e,sys)

def getDataIn2DList(file, sheetName, row, col):
    try:
        workbook=openpyxl.load_workbook(file)
        sheet=workbook[sheetName]
        rows=[] 
        for i in range(2,row+1):
            columns=[]
            for j in range(1,col+1):
                columns.append(readData(file,sheetName,i,j))  
            rows.append(columns)
        return rows
    except Exception as e:
        raise CustomException(e,sys)